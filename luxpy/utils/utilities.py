# -*- coding: utf-8 -*-
########################################################################
# <LUXPY: a Python package for lighting and color science.>
# Copyright (C) <2017>  <Kevin A.G. Smet> (ksmet1977 at gmail.com)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#########################################################################
"""
Module with utility functions and parameters
============================================

 :_PKG_PATH: absolute path to luxpy package
 
 :_SEP: operating system operator
 
 :_EPS: = 7./3 - 4./3 -1 (machine epsilon)
 
 :np2d(): Make a tuple, list or array at least a 2D numpy array.

 :np2dT(): Make a tuple, list or array at least a 2D numpy array and tranpose.

 :np3d(): Make a tuple, list or array at least a 3D numpy array.

 :np3dT(): Make a tuple, list or array at least a 3D numpy array 
           and tranpose (swap) first two axes.

 :put_args_in_db():  | Takes the args with not-None input values of a function 
                       and overwrites the values of the corresponding keys 
                       in dict db.
                     | See put_args_in_db? for more info.

 :vec_to_dict(): Convert dict to vec and vice versa.

 :getdata(): Get data from csv-file.

 :dictkv(): Easy input of of keys and values into dict 
            (both should be iterable lists).

 :OD(): Provides a nice way to create OrderedDict "literals".

 :meshblock(): | Create a meshed block.
               | (Similar to meshgrid, but axis = 0 is retained) 
               | To enable fast blockwise calculation.

 :aplit(): Split ndarray data on (default = last) axis.

 :ajoin(): Join tuple of ndarray data on (default = last) axis.

 :broadcast_shape(): | Broadcasts shapes of data to a target_shape. 
                     | Useful for block/vector calculations when numpy fails 
                       to broadcast correctly.

 :todim(): Expand x to dimensions that are broadcast-compatable 
           with shape of another array.
           
 :read_excel(): Read data from a specific Sheet and Cell_Range of an existing an Excel file.
           
 :write_excel(): Write an ndarray into specific Sheet and Cell_Range of an (existing) Excel file.
 
 :show_luxpy_tree(): Show luxpy folder structure
 
 :is_importable(): Check if a module is importable / loaded and if it doesn't exist installing it using subprocess
 
 :get_function_kwargs(): Get dictionary of a function's keyword arguments and their default values. 

 :profile_fcn(): Profile or time a function fcn.

 :unique(): Get unique elements from array.
 
 :save_pkl(): save object in pickle file
 
 :load_pkl(): load object in pickle file
 
 :imread(): read image file using imageio 
 
 :imsave(): save image file using imageio
 
 :lazy_import(): for lazy importing of a module
 
===============================================================================
"""
#------------------------------------------------------------------------------
# Package imports:
# Core:
import os
import warnings
import subprocess
import sys
import importlib
import importlib.util
import time
import cProfile
import pstats
import io
import pickle
import gzip
from collections import OrderedDict as odict
__all__ = ['odict']

# other:
import numpy as np
# import matplotlib.pyplot as plt # have become lazy imports 
# import scipy as sp # have become lazy imports 

# import lazy_loader as lazy
# __all__ += ['lazy']

#------------------------------------------------------------------------------
# os related utility parameters:
_PKG_PATH = os.path.dirname(__file__);""" Absolute path to package """ 
_PKG_PATH = _PKG_PATH[:_PKG_PATH.find("utils")-1]
_SEP = os.sep; """ Operating system separator """
__all__ += ['_PKG_PATH','_SEP']


#------------------------------------------------------------------------------
# set some general utility parameters:
_EPS = 7./3 - 4./3 -1; """ Machine epsilon """
__all__+=['_EPS']


#------------------------------------------------------------------------------
from .folder_tree import tree
__all__ += ['get_Axes3D_module',
            'np2d','np3d','np2dT','np3dT',
            'put_args_in_db','vec_to_dict',
            'loadtxt', 'savetxt','getdata',
            'dictkv','OD','meshblock','asplit','ajoin',
            'broadcast_shape','todim','read_excel', 'write_excel',
            'show_luxpy_tree', 'is_importable','get_function_kwargs',
            'profile_fcn','unique',
            'save_pkl', 'load_pkl','imread','imsave', 'lazy_import']

##############################################################################
# Start function definitions
##############################################################################
def get_Axes3D_module():
    """ Get Axes3D module from mpl_toolkits.mplot3d """
    from mpl_toolkits.mplot3d import Axes3D # lazy import
    return Axes3D

#------------------------------------------------------------------------------
def np2d(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns:
            | ndarray with .ndim >= 2
    """
    if isinstance(data, np.ndarray):# assume already atleast_2d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=2):
            return data   
        else:
            return np.atleast_2d(data)
    else:
        return np.atleast_2d(np.array(data))


def np2dT(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array and transpose.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 2 and with transposed axes.
    """
    if isinstance(data, np.ndarray):# assume already atleast_2d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=2):
            return data.T   
        else:
            return np.atleast_2d(data).T
    else:
        return np.atleast_2d(np.array(data)).T

def np3d(data):
    """
    Make a tuple, list or numpy array at least a 3d numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3
    """
    if isinstance(data, np.ndarray):# assume already atleast_3d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=3):
            return data   
        else:
            return np.expand_dims(np.atleast_2d(data),axis=0)
    else:
        return np.expand_dims(np.atleast_2d(np.array(data)),axis=0)
    
def np3dT(data): # keep last axis the same
    """
    Make a tuple, list or numpy array at least a 3d numpy array and transposed first 2 axes.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3 and with first two axes 
            | transposed (axis=3 is kept the same).
    """
    if isinstance(data,np.ndarray):# assume already atleast_3d when nd.array (user has to ensure input is an array)
        if (len(data.shape)>=3):
            return data.transpose((1,0,2))
        else:
            return np.expand_dims(np.atleast_2d(data),axis=0).transpose((1,0,2))
    else:
        return np.expand_dims(np.atleast_2d(np.aray(data)),axis=0).transpose((1,0,2))



#------------------------------------------------------------------------------
def put_args_in_db(db, args):
    """
    Takes the args with not-None input values of a function and overwrites 
    the values of the corresponding keys in dict db.
    | (args are collected with the built-in function locals(), 
    | See example usage below)
    
    Args:
        :db: 
            | dict
    
    Returns:
        :returns: 
            | dict with the values of specific keys overwritten by the 
            |      not-None values of corresponding args of a function fcn.
    
    Example usage:
        | _db = {'c' : 'c1', 'd' : 10, 'e' : {'e1':'hello', 'e2':1000}}
        |
        | def test_put_args_in_db(a, b, db = None, c = None,d = None,e = None):
        | 
        |    args = locals().copy()  # get dict with keyword input arguments to 
        |                             # function 'test_put_args_in_db'
        |     
        |     db = put_args_in_db(db,args) # overwrite non-None args in db copy.
        |     
        |    if db is not None: # unpack db for further use
        |        c,d,e = [db[x] for x in sorted(db.keys())]
        |     
        |     print(' a : {}'.format(a))
        |     print(' b : {}'.format(b))
        |     print(' db: {}'.format(db))
        |     print(' c : {}'.format(c))
        |     print(' d : {}'.format(d))
        |     print(' e : {}'.format(e))
        |     print('_db: {}'.format(_db))
 
    """
    # overwrite not-'None' input arguments in db:
    if db is not None:
        dbc = db.copy()
        dbc_keys = dbc.keys()
        for argkey in args.keys():
            if args[argkey] is not None:
                if argkey in dbc_keys:
                    dbc[argkey] =  args[argkey]
        return dbc
    else:
        return db

#------------------------------------------------------------------------------   
def vec_to_dict(vec = None, dic = None, vsize = None, keys = None):
    """
    Convert dict to vec and vice versa.
    
    Args:
        :vec: 
            | None or list or vector array, optional
        :dic: 
            | None or dict, optional
        :vsize:
            | list or vector array with size of values of dict, optional
        :keys:
            | list or vector array with keys in dict (must be provided).
        
    Returns:
        :returns:
            | x, vsize
            |   x is an array, if vec is None
            |   x is a dict, if vec is not None
    """
    if (vec is not None) & (dic is None):
        # Put values in vec in dic:
        n = 0 # keeps track of length already read from x
        dic = {}
        for i,v in enumerate(keys):
            dic[v] = vec[n + np.arange(vsize[i])]
            n += dic[v].shape[0] 
        return dic, vsize
    elif (vec is None) & (dic is not None):
        # Put values of keys in dic in vec:
        vec = []
        vsize = []
        for i,v in enumerate(keys):
            vec = np.hstack((vec, dic[v]))
            vsize.append(dic[v].shape[0])
        return vec, vsize
    else:
        raise Exception('Either vec or dict must be not None, and also not both!')

#------------------------------------------------------------------------------
def loadtxt(filename, header = None, sep = ',', dtype = float, missing_values = np.nan):
    """ 
    Load data from text file. 
    
    Args:
        :filename:
            | String with filename [+path]
        :header:
            | None, optional
            | None: no header present, 'infer' get from file.
        :sep:
            | ',', optional
            | Delimiter (',' -> csv file)
        :dtype:
            | float, optional
            | Try casting output array to this datatype.
        :missing_values:
            | np.nan, optional
            | Replace missing values with this.
            
    Returns:
        :ndarray:
            | loaded data in ndarray of type dtype or object (in case of mixed types)
    """
    with open(filename,'r') as f:
        lines = f.readlines()
    out = np.array([line.strip().split(sep) for line in lines], dtype = object)
    N = len(lines)
    if header == 'infer':
        header = out[0]
        out = out[1:]
    if dtype is not None:
        try: 
            out[out==''] = missing_values
            out = out.astype(dtype)
        except:
            out = out.astype(object)
    return out, header

def savetxt(filename, X, header = None, sep = ',', fmt = ':1.18f', aw = 'w'):
    """ 
    Save data to text file. 
    
    Args:
        :filename:
            | String with filename [+path]
        :X:
            | ndarray with data
        :header:
            | None or list, optional
            | None: no header present.
        :sep:
            | ',', optional
            | Delimiter (',' -> csv file)
        :fmt:
            | ':1.18f', optional
            | Format string for numerical data output.
            | Can be tuple/list for different output formats per column.
        :aw:
            | 'w', optional
            | options: 'w' -> write or 'a' -> append to file
    """

    if isinstance(header,list):
        header = sep.join(header)
    if fmt is None: fmt = ':g'
    if X.dtype == object:    
        lines  = [] if header is None else [header + '\n']
        for i in range(X.shape[0]):
            line = ''
            for j in range(X.shape[-1]):
                if isinstance(X[i,j],str): 
                    line = line + sep + '{:s}'.format(X[i,j])
                else:
                    fmtj = fmt[j] if isinstance(fmt,(list,tuple)) else fmt 
                    if fmtj[0] == '%': fmtj = ':' + fmtj[1:]
                    fmtj = '{' + fmtj + '}'
                    line = line + sep + fmtj.format(X[i,j])
            lines.append(line[1:] + '\n')

        with open(filename,aw) as f:
            f.writelines(lines)    
    else:
        if fmt[0] == ':': fmt = '%' + fmt[1:] 
        if header is not None:
            np.savetxt(filename, X, fmt = fmt, delimiter = sep, header = header, comments = '')
        else:
            np.savetxt(filename, X, fmt = fmt, delimiter = sep)


#------------------------------------------------------------------------------
def getdata(data, dtype = float, header = None, sep = ',', 
            datatype = 'S', copy = True, verbosity = False, missing_values = np.nan):
    """
    Get data from csv-file. 
    
    Args:
        :data: 
            | - str with path to file containing data
            | - ndarray with data
        :dtype:
            | float, optional
            | dtype of elements in ndarray data array
            | If None: mixture of datatypes is expected->dtype of output will be object
        :header:
            | None, optional
            |   - None: no header in file
            |   - 'infer': infer headers from file
        :sep:
            | ',' or '\t' or other char, optional
            | Column separator in data file
        :datatype':
            | 'S',optional 
            | Specifies a type of data. 
            | Is used when creating column headers (:column: is None).
            |   -'S': light source spectrum
            |   -'R': reflectance spectrum
            |   or other.   
        :copy:
            | True, optional
            | Return a copy of ndarray 
        :verbosity:
            | True, False, optional
            | Print warning when inferring headers from file.
    
    Returns:
        :returns:
            | data as ndarray 
      
    """
    if isinstance(data,str):
        datafile = data
        input_is_string = True
        if header == 'infer':
            #data = (np.array((np.genfromtxt(datafile, delimiter = sep, dtype = None,skip_header = 1, encoding = None)).tolist(),dtype=object))
            #header = np.genfromtxt(datafile, delimiter = sep, dtype = str, max_rows = 1)
            data, header = loadtxt(datafile, sep = sep, dtype = dtype, header = header, missing_values = missing_values)
            if verbosity == True:
                warnings.warn('getdata(): Infering HEADERS from data file: {}!'.format(datafile))
        else:
            #data = (np.array((np.genfromtxt(datafile, delimiter = sep, dtype = None, encoding = None)).tolist(),dtype=object))
            data, header = loadtxt(datafile, sep = sep, dtype = dtype, header = header, missing_values = missing_values)
    
    else:
        input_is_string = False
        
    if copy == True: data = data.copy()
    if (dtype is not None) & (input_is_string): 
        try: 
            data = data.astype(dtype)
        except: 
            pass
    return data

#--------------------------------------------------------------------------------------------------
def dictkv(keys=None,values=None, ordered = True): 
    """
    Easy input of of keys and values into dict.
    
    Args:
        :keys: 
            | iterable list[str,...] of keys
        :values:
            | iterable list[...,..., ] of values
        :ordered:
            | True, False, optional
            | True: creates an ordered dict using 'collections.OrderedDict()'
    Returns:
        :returns:
            | (ordered) dict
    """
    if ordered is True:
        return odict(zip(keys,values))
    else:
        return dict(zip(keys,values))
    
class OD(object):
    """
    This class provides a nice way to create OrderedDict "literals".
    
    """
    def __getitem__(self, slices):
        if not isinstance(slices, tuple):
            slices = slices,
        return odict((slice.start, slice.stop) for slice in slices)
# Create a single instance; we don't ever need to refer to the class.
OD = OD()

#--------------------------------------------------------------------------------------------------
def meshblock(x,y):
    """
    Create a meshed block from x and y.
    
    | (Similar to meshgrid, but axis = 0 is retained).
    | To enable fast blockwise calculation.
    
    Args: 
        :x: 
            | ndarray with ndim == 2
        :y: 
            | ndarray with ndim == 2
        
    Returns:
        :X,Y: 
            | 2 ndarrays with ndim == 3 
            |   X.shape = (x.shape[0],y.shape[0],x.shape[1])
            |   Y.shape = (x.shape[0],y.shape[0],y.shape[1])
    """
    Y = np.transpose(np.repeat(y,x.shape[0],axis=0).reshape((y.shape[0],x.shape[0],y.shape[1])),axes = [1,0,2])
    X = np.repeat(x,y.shape[0],axis=0).reshape((x.shape[0],y.shape[0],x.shape[1]))
    return X,Y

#---------------------------------------------------------------------------------------------------
def asplit(data):
    """
    Split data on last axis
    
    Args:
        :data: 
            | ndarray
        
    Returns:
        :returns: 
            | ndarray, ndarray, ... 
            |   (number of returns is equal data.shape[-1])
    """
    #return np.array([data.take([x],axis = len(data.shape)-1) for x in range(data.shape[-1])])
    return [data[...,x] for x in range(data.shape[-1])]


def ajoin(data):
    """
    Join data on last axis.
    
    Args:
        :data:
            | tuple (ndarray, ndarray, ...)
        
    Returns:
        :returns:
            | ndarray (shape[-1] is equal to tuple length)
    """
    if data[0].ndim == 2: #faster implementation
        return np.transpose(np.concatenate(data,axis=0).reshape((np.hstack((len(data),data[0].shape)))),(1,2,0))
    elif data[0].ndim == 1:
        return np.concatenate(data,axis=0).reshape((np.hstack((len(data),data[0].shape)))).T
    elif data[0].ndim == 3:
        return np.dstack(data)
    else:
        return np.hstack(data)[0]
    

#---------------------------------------------------------------------------------------------------
def broadcast_shape(data,target_shape = None, expand_2d_to_3d = None, axis0_repeats = None, axis1_repeats = None):
    """
    Broadcasts shapes of data to a target_shape.
    
    | Useful for block/vector calc. when numpy fails to broadcast correctly.
    
    Args:
        :data: 
            | ndarray 
        :target_shape: 
            | None or tuple with requested shape, optional
            |   - None: returns unchanged :data:
        :expand_2d_to_3d:
            | None (do nothing) or ..., optional 
            | If ndim == 2, expand from 2 to 3 dimensions
        :axis0_repeats:
            | None or number of times to repeat axis=0, optional
            |   - None: keep axis=0 same size
        :axis1_repeats:
            | None or number of times to repeat axis=1, optional
            |   - None: keep axis=1 same size

    Returns:
        :returns: 
            | reshaped ndarray
    """
    data = np2d(data)
    
    # expand shape along axis (useful for some functions that allow block-calculations when the data is only 2d )
    if (expand_2d_to_3d is not None) & (len(data.shape) == 2):
        data = np.expand_dims(data, axis = expand_2d_to_3d)
     
    if target_shape is not None:
        dshape = data.shape
        if dshape != target_shape:
            axis_of_v3 = len(target_shape)-1
            if (dshape[0] != target_shape[0]): # repeat along axis 0
                if axis0_repeats is None:
                    axis0_repeats = (target_shape[0]-dshape[0] + 1)
                data = np.repeat(data,axis0_repeats,axis = 0)
            if (len(target_shape)>2) & (len(data.shape)==2): # repeat along axis 1, create axis if necessary
                data = np.expand_dims(data,axis = axis_of_v3-1) # axis creation
                dshape = data.shape
                if (dshape[1] != target_shape[1]):
                    if axis1_repeats is None:
                        axis1_repeats = (target_shape[1]-dshape[1] + 1) # repititon
                    data = np.repeat(data,axis1_repeats,axis = 1)

        for i in range(2):
            if (data.shape[i] > 1) & (data.shape[i] != target_shape[i]):
                raise Exception('broadcast_shape(): Cannot match dim of data with target: data.shape[i]>1  & ...  != target.shape[i]')
    return data


def todim(x,tshape, add_axis = 1, equal_shape = False): 
    """
    Expand x to dims that are broadcast-compatable with shape of another array.
    
    Args:
        :x: 
            | ndarray
        :tshape: 
            | tuple with target shape
        :add_axis:
            | 1, optional
            | Determines where in x.shape an axis should be added
        :equal_shape:
            | False or True, optional
            | True: expand :x: to identical dimensions (speficied by :tshape:)
            
    Returns:
        :returns:
            | ndarray broadcast-compatable with tshape.
    """
    if x is None:
        return np.broadcast_arrays(x,np.ones(tshape))[0]
    else:
        x = np2d(x)
        sx = x.shape
        lsx = len(sx)
        ltshape = len(tshape)
        if (sx == tshape):
            pass
        else:
            
            if ((lsx == 1) | (sx == (1,tshape[-1])) | (sx == (tshape[-1],1))): 
                if (sx == (tshape[-1],1)):
                    x = x.T
                if lsx != ltshape:
                    x = np.expand_dims(x,0)
            elif (lsx == 2):
                if (ltshape == 3):
                    sd = np.setdiff1d(tshape, sx,assume_unique=True)
                    if len(sd) == 0:
                        ax = add_axis
                    else:
                        ax = np.where(tshape==sd)[0][0]
                    x = np.expand_dims(x,ax)
                else:
                    raise Exception("todim(x,tshape): dimensions do not match for 2d arrays.")  
            else:
                raise Exception("todim(x,tshape): no matching dimensions between 3d x and tshape.")
        if equal_shape == False:
            return x
        else:
            return np.ones(tshape)*x #make dims of x equal to those of a (tshape)
        
#------------------------------------------------------------------------------
def read_excel(filename, sheet_name = None, cell_range = None, dtype = float, 
               force_dictoutput = False, out = 'X'):
    """
    Read excel file using openpyxl.

    Args:
        :filename:
            | string with [path/]filename of Excel file.
        :sheet_name:
            | None, optional
            | If None: read all sheets
            | If string or tuple/list of strings: read these sheets.
        :cell_range:
            | None, optional
            | Read all data on sheet(s).
            | If string range (e.g. 'B2:C4') or tuple/list of cell_ranges: read this range.
            | If tuple/list: then length must match that of the list of sheet_names!
        :dtype:
            | float, optional
            | Try to cast the output data array(s) to this type. In case of failure, 
            | data type will be 'object'.
        :force_dictoutput:
            | False, optional
            | If True: output will always be a dictionary (sheet_names are keys) 
            |          with the requested data arrays.
            | If False: in case only a single sheet_name is supplied or only a single
            |          sheet is present, then the output will be an ndarray!
        :out:
            | 'X', optional
            | String specifying requested output (eg. 'X' or 'X,wb' with wb the loaded workbook)
            
    Returns:
        :X:
            | dict or ndarray (single sheet and force_dictoutput==False) 
            | with data in requested ranges. 
        :wb:
            | If in :out: the loaded workbook is also output.
    """
    
    success = is_importable('openpyxl')
    if success:
        try:
            import openpyxl # lazy import
        except:
            raise Exception("Could not import (nor pip install openpyxl)! Please try a manual install. And retry.")
            
    wb = openpyxl.load_workbook(filename = filename, data_only=True)
    
    # process sheet_names:
    if sheet_name is None:
        sheet_names = wb.sheetnames
        single_sheet = True if len(sheet_names) == 1 else False 
    else:
        if isinstance(sheet_name,str): 
            sheet_names = [sheet_name]
            single_sheet = True
        else:
            sheet_names = sheet_name
            single_sheet = True if len(sheet_names) == 1 else False 
    if force_dictoutput: single_sheet = False
            
    # process cell_ranges:
    if cell_range is not None: 
        if isinstance(cell_range,str): 
            cell_ranges = [cell_range]*len(sheet_names)
        else:
            cell_ranges = cell_range
            if len(sheet_names) != len(cell_ranges):
                raise Exception("Number of cell_ranges doesn't match number of sheet_names")
    else:
        cell_ranges = [None]*len(sheet_names)
    
    
    # loop of sheet_names and cell_ranges:
    X = {}
    for sheet_name, cell_range in zip(sheet_names,cell_ranges):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if cell_range is not None:
                if ':' in cell_range:
                    left = cell_range[:cell_range.index(':')]
                    right = cell_range[(cell_range.index(':')+1):]
                else:
                    left = cell_range
                    right = None
                left = openpyxl.utils.cell.coordinate_to_tuple(left)
                right = openpyxl.utils.cell.coordinate_to_tuple(right) if right is not None else None
                
                nrows = 1 if right is None else (right[0] - left[0]) + 1
                ncols = 1 if right is None else (right[1] - left[1]) + 1
            else:
                nrows, ncols = 0, 0
                for row in sheet.rows:
                    nrows += 1
                    for cell in row:
                        ncols += 1
                left = (1,1)
                
            X[sheet_name] = np.empty((nrows,ncols),dtype = object)
            for i in range(nrows):
                row = left[0] + i
                for j in range(ncols):
                    col = left[1] + j
                    X[sheet_name][i,j] = sheet.cell(row,col).value
                    
            # get rid of extra None cols:
            if cell_range is None: 
                x = X[sheet_name]
                X[sheet_name] = x[:,:-np.hstack((1,np.diff((((x != None)*1).sum(0)!=0).cumsum())))[::-1].argmax()]
                
            try: 
                X[sheet_name] = X[sheet_name].astype(dtype)
            except:
                pass
            
    if single_sheet: X = X[sheet_name]
    
    if out == 'X': 
        return X
    elif out == 'X,wb':
        return X,wb
    else:
        return eval(out)

def write_excel(filename, X, sheet_name = None, cell_range = None):
    """
    Write data to an excel file using openpyxl.
    
    Args:
        :filename:
            | string with [path/]filename of Excel file.
        :sheet_name:
            | None, optional
            | If None: use first one (or the keys in :X: when it is a dictionary)
            | If string: use this sheet.
            | If tuple/list of strings: use these to write the data in :X: (if :X: is a list/tuple of ndarrays)
        :X:
            | ndarray, list/tuple or dict
            | If ndarray/list/tuple: sheet_names must be supplied explicitely in :sheet_names:
            | If dict: keys must be sheet_names        
        :cell_range:
            | None, optional
            | Read all data on sheet(s).
            | If string range (e.g. 'B2:C4') or tuple/list of cell_ranges: read this range.
            | If tuple/list: then length must match that of the list of sheet_names!
    """
    
    success = is_importable('openpyxl')
    if success:
        try:
            import openpyxl # lazy import
        except:
            raise Exception("Could not import (nor pip install openpyxl)! Please try a manual install. And retry.")

    if os.path.exists(filename):
        #wb = openpyxl.load_workbook(filename = filename, data_only = True)

        # Get pre-existing data from workbook
        X0, wb = read_excel(filename, sheet_name = None, cell_range = None, dtype = float, force_dictoutput = True, out = 'X,wb')
        
    else:
        head_tail = os.path.split(filename)
        if not os.path.exists(head_tail[0]):
            os.makedirs(head_tail[0], exist_ok = True)
        wb = openpyxl.Workbook()
        
    # Get number of data arrays to write: 
    if isinstance(X, np.ndarray): X = [X] 
    N = len(X) 

    # process sheet_names:
    if sheet_name is None:
        if N == 1: 
            sheet_names = wb.sheetnames
            sheet_names = [sheet_names[0]] # use first sheet
        else:
            if isinstance(X,dict):
                sheet_names = list(X.keys())
            else:
                raise Exception('List/tuple of ndarrays as input, but no sheet_names provided!')
    else:
        if isinstance(sheet_name,str): 
            sheet_names = [sheet_name]
        else:
            sheet_names = sheet_name
    if len(sheet_names) != N:
        raise Exception("Number of sheet_names doesn't match number of supplied data arrays.")

    # process cell_ranges:
    if cell_range is not None: 
        if isinstance(cell_range,str): 
            cell_ranges = [cell_range]*N
        else:
            cell_ranges = cell_range
            if len(cell_ranges) != N:
                raise Exception("Number of cell_ranges doesn't match number of supplied data arrays")
    else:
        cell_ranges = [None]*N
    
    # prepare X -> convert to dict:
    X = dict(zip(sheet_names,X))

    # loop of sheet_names and cell_ranges:
    for sheet_name, cell_range  in zip(sheet_names,cell_ranges):
        x = X[sheet_name] # get data to write
        
        if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        
        if cell_range is not None:
            if ':' in cell_range:
                left = cell_range[:cell_range.index(':')]
                right = cell_range[(cell_range.index(':')+1):]
            else:
                left = cell_range
                right = None
            left = openpyxl.utils.cell.coordinate_to_tuple(left)
            right = openpyxl.utils.cell.coordinate_to_tuple(right) if right is not None else None
            
            nrows = 1 if right is None else (right[0] - left[0]) + 1
            ncols = 1 if right is None else (right[1] - left[1]) + 1
        else:
            nrows, ncols = 0, 0
            for row in sheet.rows:
                nrows += 1
                for cell in row:
                    ncols += 1
            left = (1,1)
                
        for i in range(nrows):
            row = left[0] + i
            for j in range(ncols):
                col = left[1] + j
                sheet.cell(row, col, value = x[i,j])
                
    wb.save(filename = filename)
    
    return None

    
def show_luxpy_tree(omit = ['.pyc','__pycache__',
                            '.txt','.dat','.csv','.npz',
                            '.png','.jpg','.md','.pdf','.ini','.log', '.rar',
                            'drivers','SDK_','dll','bak']):
    """
    Show luxpy foler tree.
    
    Args:
        :omit:
            | List of folders and file-extensions to omit.
            
    Returns:
        None
    """
    tree(_PKG_PATH, omit = omit)
    return None


#------------------------------------------------------------------------------
def is_importable(string, pip_string = None, try_pip_install = False):
    """
    Check if string is importable/loadable. If it doesn't then try to 'pip install' it using subprocess.
    Returns None if succesful, otherwise throws and error or outputs False.
    
    Args:
        :string:
            | string with package or module name
        :pip_string:
            | string with package or module name as known by pip
            | If None: use the import string
        :try_pip_install:
            | False, optional
            | True: try pip installing it using subprocess
    
    Returns:
        :success:
            | True if importable, False if not.
    """ 
    success = importlib.util.find_spec(string) is not None
    if (not success) & (try_pip_install == True): 
        if pip_string is None: pip_string = string
        try:
            print("Trying to 'pip install {:s}' using subprocess.".format(pip_string))
            success = subprocess.call(["pip", "install", "{:s}".format(pip_string)])
            print("subprocess output: ", success)
            if success != 0:
                raise Exception("Tried importing '{:s}', then tried pip installing it. Please install it manually: pip install {:s}".format(string,pip_string))  
            else:
                print("'pip install {:s}' succesful".format(pip_string))
            success = importlib.util.find_spec(string) is not None
        except:
            success = False
            raise Exception("Tried importing '{:s}', then tried pip installing it. Please install it manually: pip install {:s}".format(string,pip_string))   
    return success    
#------------------------------------------------------------------------------
def get_function_kwargs(f):
    """
    Get dictionary of a function's keyword arguments and their default values. 
    
    Args:
        :f:
            | function name
    
    Returns:
        :dict:
            | Dict with the function's keyword arguments and their default values
            | Is empty if there are no defaults (i.e. f.__defaults__ or f.__kwdefaults__ are None).
    """
    kwdefs = f.__kwdefaults__
    if kwdefs is None: kwdefs = {}
    names = f.__code__.co_varnames[:f.__code__.co_argcount]
    if f.__defaults__ is not None:
        d = dict(zip(names[::-1][:len(f.__defaults__)][::-1],f.__defaults__))
    else:
        d = {}
    d.update(kwdefs)
    return d

#------------------------------------------------------------------------------
def profile_fcn(fcn, profile=True, sort_stats = 'tottime', output_file=None):
    """
    Profile or time a function fcn.
    
    Args:
        :fcn:
            | function to be profiled or timed (using time.time() difference)
        :profile:
            | True, optional
            | Profile the function, otherwise only time it.
        :sort_stats:
            | 'tottime', optional
            | Sort profile results according to sort_stats ('tottime', 'cumtime',...)
        :output_file:
            | None, optional
            | If not None: output result to output_file.
            
    Return:
        :ps:
            | Profiler output
            
    """
   
    if profile == False:
        start = time.time()
        fcn()
        dt = time.time()-start
        print('%s %f' % ('Time elapsed: ', dt))
        return dt
    else:
        pr = cProfile.Profile()
        pr.enable()
        fcn()
        pr.disable()
        s = io.StringIO()
        ps = pstats.Stats(pr, stream=s).sort_stats(sort_stats)
        ps.print_stats()
        if output_file is not None:
            with open(output_file, 'w+') as f:
                f.write(s.getvalue())
        return ps
    
#------------------------------------------------------------------------------
def unique(array, sort = True):
    """ 
    Get unique elements from array.
    
    Args:
        :array:
            | array to get unique elements from.
        :sort:
            | True, optional
            | If True: get sorted unique elements.
            
    Returns:
        :unique_array:
            | ndarray with (sorted) unique elements.
    """
    if sort:
        return np.unique(array) 
    else:
        uniq, index = np.unique(array, return_index=True)
        return uniq[index.argsort()]

#------------------------------------------------------------------------------
def save_pkl(filename, obj, compresslevel = 0): 
    """ 
    Save an object in a (gzipped) pickle file.
    
    Args:
        :filename:
            | str with filename of pickle file.
        :obj:
            | python object to save
        :compresslevel:
            | 0, optional
            | If > 0: use gzip to compress pkl file.
    
    Returns:
        :None:
    """
    _open = (lambda file,w: gzip.open(file+'.gz', w, compresslevel = compresslevel)) if (compresslevel > 0) else (lambda file, w: open(file,w))
    with _open(filename, 'wb') as handle:
        pickle.dump(obj, handle, protocol=pickle.HIGHEST_PROTOCOL)
        
def load_pkl(filename, gzipped = False):
    """ 
    Load the object in a (gzipped) pickle file.
    
    Args:
        :filename:
            | str with filename of pickle file.
        :gzipped:
            | False, optional 
            | If True: '.gz' will be added to filename before opening. 
        
    Returns:
        :obj:
            | loaded python object
    """
    obj = None
    if gzipped and (filename[-3:] != '.gz'): filename = filename + '.gz'
    _open = gzip.open if filename[-3:] == '.gz' else open
    with _open(filename, 'rb') as handle:
        obj = pickle.load(handle)
    return obj

#------------------------------------------------------------------------------
def _try_imageio_import(use_freeimage=True): 
    success = is_importable('imageio', try_pip_install = True)
    if success: 
        import imageio # lazy import
    else:
        imageio = None
    try: 
        if use_freeimage: imageio.plugins.freeimage.download() 
    except:
        if use_freeimage: 
            print("!!!      imageio.plugins.freeimage.download() failed. !!!")
            print("                                                          ")
            print("  Try installing the freeimage plugin manually.") 
            print("  or, try downgrading imageio")
            print("  or, wait until the developers of imageio fix this.")
            print("  or, try saving it using PIL or opencv or openexc or other.")
    return imageio

def imsave(file, img, use_freeimage = False):
    """ Save image using imageio"""
    imageio = _try_imageio_import(use_freeimage) # lazy-import
    try: 
        imageio.v3.imwrite(file, img)
    except:
        imageio.imwrite(file, img)
    
def imread(file, use_freeimage = False):
    """ Read image using imageio"""
    imageio = _try_imageio_import(use_freeimage) # lazy-import
    try: 
        return imageio.v3.imread(file)
    except:
        return imageio.imread(file)

#------------------------------------------------------------------------------
def lazy_import(name):
    """ Lazy import of module """
    spec = importlib.util.find_spec(name)
    loader = importlib.util.LazyLoader(spec.loader)
    spec.loader = loader 
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    loader.exec_module(module)
    return module